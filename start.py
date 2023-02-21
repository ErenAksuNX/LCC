import openpyxl as xl
import datetime


#  gibt zurück ob der zeile
def ist_gueltiger_zeitraum(r_m, q_m, r_jahr, q_jahr) -> bool:
    if (r_m == 12 and q_m == 1) or q_m == r_m:
        return True
    elif (q_m + 1) % 12 == r_m:
        return True
    else:
        return False


def ist_Gueltige_Zeile(r_frz, q_frz, r_dauer, q_dauer, r_werk, q_werk):
    if r_frz in q_frz and r_dauer == q_dauer and r_werk in q_werk:
        return True
    else:
        return False


def aufrunden(n):
    return float("{:.2f}".format(n))


def run(quma='C:/Users/eaksu/OneDrive - National Express Rail GmbH/Desktop/Projekte/Python Projekte/LCC/_Quma.xlsx',
        rechnung='C:/Users/eaksu/OneDrive - National Express Rail GmbH/Desktop/Projekte/Python Projekte/LCC/Rechnung.xlsx',
        speicherort=""):
    # Workbooks laden
    wb_quma = xl.load_workbook(quma)
    wb_rechnung = xl.load_workbook(rechnung)

    # Worksheets Auswählen
    ws_quma = wb_quma.active
    ws_rechnung = wb_rechnung.active

    ws_151 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\151.XLSX").active
    ws_152 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\152.XLSX").active
    ws_153 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\153.XLSX").active
    ws_154 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\154.XLSX").active
    ws_155 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\155.XLSX").active
    ws_156 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\156.XLSX").active
    ws_157 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\157.XLSX").active
    ws_158 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\158.XLSX").active
    ws_159 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\159.XLSX").active
    ws_160 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\160.XLSX").active
    ws_351 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\351.XLSX").active
    ws_352 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\352.XLSX").active
    ws_353 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\353.XLSX").active
    ws_354 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\354.XLSX").active
    ws_355 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\355.XLSX").active
    ws_356 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\356.XLSX").active
    ws_357 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\357.XLSX").active
    ws_358 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\358.XLSX").active
    ws_359 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\359.XLSX").active
    ws_360 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\360.XLSX").active
    ws_362 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\362.XLSX").active
    ws_363 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\363.XLSX").active
    ws_364 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\364.XLSX").active
    ws_365 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\365.XLSX").active
    ws_366 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\366.XLSX").active
    ws_367 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\367.XLSX").active
    ws_368 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\368.XLSX").active
    ws_369 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\369.XLSX").active
    ws_370 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\370.XLSX").active
    ws_371 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\371.XLSX").active
    ws_372 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\372.XLSX").active
    ws_373 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\373.XLSX").active
    ws_374 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\374.XLSX").active
    ws_375 = xl.load_workbook(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Projekte\Python Projekte\LCC\qfrz\375.XLSX").active

    avnrs = []
    count = 1
    # Pro Zeile in der Rechnungsdatei in der Quma Matrix suchen
    for r_zeile in ws_rechnung.iter_rows(values_only=True):

        count += 1

        # Ungültige und die erste Zeilen werden übersprungen
        if r_zeile[18] == "x" or r_zeile[18] == "X" or r_zeile[0] == "Geschäftsjahr":
            continue

        # Datum Werk und Fahrzeugnummer aus der Rechnungszeile speichern
        datum: datetime.datetime = r_zeile[2]
        try:
            r_monat = datum.month
            r_jahr = datum.year
        except:
            continue
        r_werk = r_zeile[19]
        r_frz = r_zeile[20]
        r_dauer = r_zeile[10]
        typo = r_zeile[18]

        if typo is not None:
            continue

        # In dieser Schleife wird in der Quma Matrix gesucht
        for q_zeile in ws_quma.iter_rows(values_only=True):

            if q_zeile[32] == "Monat" or q_zeile[1] in avnrs:
                continue

            # Das Datum aus der Quma Zeile speichern
            q_monat = int(q_zeile[32])
            q_jahr = int(q_zeile[33])
            # avnr ist der Key der in die rechnungsdatei soll
            avnr = q_zeile[1]
            # q_werk ist das Werk welches in der Quma Matrix ist
            q_werk = q_zeile[13]
            # q_frz ist das Fahrzeug um welches es sich in der Quma Matrix handelt
            q_frz = q_zeile[0]

            # Ungültige Spalten werden übersprungen

            if not (r_monat == 12 or r_monat == 11) and q_monat <= (r_monat + 2) % 12:
                break

            # hier wird die Dauer der Arbeit aus der Quma Matrix berechnet

            q_dauer = aufrunden(float(q_zeile[11]) / 60)

            if "366" in q_frz and q_zeile[11] == 12 and r_dauer == 0.2 and r_frz == "366":
                print(1)

            if not ist_Gueltige_Zeile(r_frz, q_frz, r_dauer, q_dauer, r_werk, q_werk):
                continue

            if ist_gueltiger_zeitraum(r_monat, q_monat, r_jahr, q_jahr):

                ws_rechnung.cell(row=count, column=22).value = avnr
                avnrs.append(avnr)
                try:
                    print(avnr)
                except:
                    pass
                break

    wb_rechnung.save("export.xlsx")
